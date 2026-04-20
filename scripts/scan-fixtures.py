#!/usr/bin/env python3
"""Scan every xlsx fixture cell for a literal string.

Usage: uv run python scripts/scan-fixtures.py <needle>
Exit code: 0 if no matches, 1 if any match (CI-friendly).
"""

import pathlib
import sys

import openpyxl

if len(sys.argv) != 2:
    print("usage: scan-fixtures.py <needle>", file=sys.stderr)
    sys.exit(2)

needle = sys.argv[1]
repo_root = pathlib.Path(__file__).resolve().parent.parent
fixtures_dir = repo_root / "test" / "fixtures"

hits = 0
for xlsx in sorted(fixtures_dir.glob("*.xlsx")):
    wb = openpyxl.load_workbook(xlsx, data_only=False)
    for sheet_name in wb.sheetnames:
        for row in wb[sheet_name].iter_rows(values_only=True):
            if any(needle in str(cell) for cell in row if cell is not None):
                print(f"{xlsx.name} [{sheet_name}]: {row}")
                hits += 1

if hits == 0:
    print(f"clean — no fixture cell contains {needle!r}")
    sys.exit(0)
print(f"{hits} hit(s)")
sys.exit(1)

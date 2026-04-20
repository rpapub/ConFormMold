"""
Generate test fixture .xlsx files for ConFormMold.
Run with: uv run generate_fixtures.py
"""

import datetime
import pathlib
import openpyxl
from openpyxl.styles import Font

OUTPUT_DIR = pathlib.Path(__file__).parent

HEADER_STANDARD       = ["Name", "Value", "Description"]
HEADER_STANDARD_TYPED = ["Name", "Value", "Description", "DataType"]
HEADER_ASSET          = ["Name", "Asset", "OrchestratorAssetFolder", "Description"]
HEADER_ASSET_TYPED    = ["Name", "Asset", "OrchestratorAssetFolder", "Description", "ValueType"]


def write_sheet(ws, header, rows):
    ws.append(header)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for row in rows:
        ws.append(row)


# ---------------------------------------------------------------------------
# Config_Basic.xlsx — strings and ints only, no `using System;` expected
# ---------------------------------------------------------------------------
def make_basic():
    wb = openpyxl.Workbook()

    ws_settings = wb.active
    ws_settings.title = "Settings"
    write_sheet(ws_settings, HEADER_STANDARD, [
        ["OrchestratorQueueName", "BasicQueue",      "Orchestrator queue name."],
        ["OrchestratorFolderPath", "RPA/Basic",      "Orchestrator folder path."],
        ["MaxItemsPerRun",         10,                "Max items to process per run."],
        ["RetryCount",             3,                 "Number of retries on failure."],
        ["LogPrefix",             "BASIC",            "Prefix for log messages."],
    ])

    ws_constants = wb.create_sheet("Constants")
    write_sheet(ws_constants, HEADER_STANDARD, [
        ["MaxRetryNumber",              0,     "Must be 0 when using Orchestrator queues."],
        ["MaxConsecutiveSystemExceptions", 3,  "Stop job after this many consecutive errors."],
        ["RetryNumberGetTransactionItem",  2,  "Retries for GetTransactionItem activity."],
    ])

    ws_assets = wb.create_sheet("Assets")
    write_sheet(ws_assets, HEADER_ASSET, [])  # empty — empty class expected

    wb.save(OUTPUT_DIR / "Config_Basic.xlsx")
    print("Created Config_Basic.xlsx")


# ---------------------------------------------------------------------------
# Config_Types.xlsx — all supported C# types, `using System;` expected
# Covers: string, int, double, bool, DateOnly, DateTime, TimeOnly,
#         OrchestratorAsset (untyped + string/int/bool),
#         DataType=credential, DataType=asset, _TargetType directive
# ---------------------------------------------------------------------------
def make_types():
    wb = openpyxl.Workbook()

    ws_settings = wb.active
    ws_settings.title = "Settings"
    write_sheet(ws_settings, HEADER_STANDARD_TYPED, [
        # Name            Value                                  Description                               DataType
        ["FeatureName",   "TypesDemo",                           "string",                                 None],
        ["MaxItems",      42,                                    "int",                                    None],
        ["Threshold",     3.14,                                  "double",                                 None],
        ["IsEnabled",     True,                                  "bool",                                   None],
        ["CutoffDate",    datetime.date(2025, 12, 31),           "DateOnly — date only, time is 00:00:00", None],
        ["ScheduledAt",   datetime.datetime(2025, 6, 15, 9, 30), "DateTime — has time component",          None],
        ["DailyRunTime",  datetime.time(8, 0, 0),                "TimeOnly — time only, no date",          None],
        ["SapCredential", "Default/cred_sap_types",              "SAP Orchestrator credential.",           "credential"],
        ["QueueNameRef",  "Default/cfgtree_queue_types",         "Queue name Orchestrator asset.",         "asset"],
    ])

    ws_constants = wb.create_sheet("Constants")
    write_sheet(ws_constants, HEADER_STANDARD, [
        ["Pi",            3.14159,                              "double — mathematical constant"],
        ["MaxRetryNumber", 0,                                   "int"],
        ["StrictMode",    False,                                "bool"],
        ["ExpiresOn",     datetime.date(2026, 1, 1),            "DateOnly"],
        ["CreatedAt",     datetime.datetime(2024, 3, 1, 12, 0), "DateTime"],
        ["WindowOpen",    datetime.time(9, 0, 0),               "TimeOnly"],
        ["WindowClose",   datetime.time(17, 30, 0),             "TimeOnly"],
    ])

    ws_assets = wb.create_sheet("Assets")
    write_sheet(ws_assets, HEADER_ASSET_TYPED, [
        # Name               Asset name                       Folder             Description                    ValueType
        ["CredentialM365",   "cred_m365_types",               "Shared",          "M365 service credential.",    None],
        ["CredentialFtp",    "cred_ftp_types",                "Shared",          "FTP server credential.",      None],
        ["QueueNameAsset",   "cfgtree_queue_name_types",      "ConFigTree/Test", "Input queue name.",           "string"],
        ["MaxItemsAsset",    "cfgtree_max_items_types",       "ConFigTree/Test", "Max items to process.",       "int"],
        ["StrictModeAsset",  "cfgtree_strict_mode_types",     "ConFigTree/Test", "Strict processing toggle.",   "bool"],
    ])

    ws_targets = wb.create_sheet("Targets")
    write_sheet(ws_targets, HEADER_STANDARD, [
        ["_TargetType", "Acme.External.TargetConfig",   None],
        ["Host",        "targets.example.com",          "Target host address."],
        ["Port",        8080,                           "Target port number."],
        ["Enabled",     True,                           "Enable target connection."],
    ])

    wb.save(OUTPUT_DIR / "Config_Types.xlsx")
    print("Created Config_Types.xlsx")


# ---------------------------------------------------------------------------
# Config_Types.toml — all supported C# types for TOML input
# Covers: string, int, double, bool, DateOnly, DateTime, TimeOnly,
#         OrchestratorAsset (untyped + string/int/bool), _TargetType directive
# (DataType=credential/asset are xlsx-only — no TOML parser equivalent)
# ---------------------------------------------------------------------------
def make_types_toml():
    content = """\
# Config_Types.toml — all supported C# types for TOML input
# Covers: string, int, double, bool, DateOnly, DateTime, TimeOnly,
#         OrchestratorAsset (untyped + string/int/bool), _TargetType directive

[Settings]
FeatureName  = "TypesDemo"
MaxItems     = 42
Threshold    = 3.14
IsEnabled    = true
CutoffDate   = 2025-12-31
ScheduledAt  = 2025-06-15T09:30:00
DailyRunTime = 08:00:00

[Constants]
Pi             = 3.14159
MaxRetryNumber = 0
StrictMode     = false
ExpiresOn      = 2026-01-01
CreatedAt      = 2024-03-01T12:00:00
WindowOpen     = 09:00:00
WindowClose    = 17:30:00

[Assets]
CredentialM365  = { assetName = "cred_m365_types",          folder = "Shared",          description = "M365 service credential." }
CredentialFtp   = { assetName = "cred_ftp_types",           folder = "Shared",          description = "FTP server credential." }
QueueNameAsset  = { assetName = "cfgtree_queue_name_types", folder = "ConFigTree/Test", description = "Input queue name.",        valueType = "string" }
MaxItemsAsset   = { assetName = "cfgtree_max_items_types",  folder = "ConFigTree/Test", description = "Max items to process.",    valueType = "int" }
StrictModeAsset = { assetName = "cfgtree_strict_mode_types",folder = "ConFigTree/Test", description = "Strict processing toggle.", valueType = "bool" }

[Targets]
_TargetType = "Acme.External.TargetConfig"
Host    = "targets.example.com"
Port    = 8080
Enabled = true
"""
    (OUTPUT_DIR / "Config_Types.toml").write_text(content, encoding="utf-8")
    print("Created Config_Types.toml")


# ---------------------------------------------------------------------------
# Config_Assets.xlsx — focus on asset schema and OrchestratorAsset output
# ---------------------------------------------------------------------------
def make_assets():
    wb = openpyxl.Workbook()

    ws_settings = wb.active
    ws_settings.title = "Settings"
    write_sheet(ws_settings, HEADER_STANDARD, [
        ["Environment",  "UAT",   "Deployment environment."],
        ["LogLevel",     "Info",  "Minimum log level."],
    ])

    ws_constants = wb.create_sheet("Constants")
    write_sheet(ws_constants, HEADER_STANDARD, [
        ["MaxRetryNumber", 0, "Must be 0 when using Orchestrator queues."],
    ])

    ws_assets = wb.create_sheet("Assets")
    write_sheet(ws_assets, HEADER_ASSET, [
        ["CredentialSap",     "cred_sap_uat",       "SAP",     "SAP system credential."],
        ["CredentialM365",    "cred_m365_uat",      "Shared",  "M365 credential."],
        ["CredentialFtp",     "cred_ftp_uat",       "FTP",     "FTP server credential."],
        ["ApiKeyPayment",     "apikey_payment_uat", "API",     "Payment gateway API key."],
    ])

    wb.save(OUTPUT_DIR / "Config_Assets.xlsx")
    print("Created Config_Assets.xlsx")


# ---------------------------------------------------------------------------
# Config_MultiSheet.xlsx — 5 sheets, tests dynamic sheet detection (#18)
# ---------------------------------------------------------------------------
def make_multi_sheet():
    wb = openpyxl.Workbook()

    ws_settings = wb.active
    ws_settings.title = "Settings"
    write_sheet(ws_settings, HEADER_STANDARD, [
        ["OrchestratorQueueName", "MultiQueue", "Orchestrator queue name."],
        ["MaxItemsPerRun",        50,            "Max items per run."],
    ])

    ws_constants = wb.create_sheet("Constants")
    write_sheet(ws_constants, HEADER_STANDARD, [
        ["MaxRetryNumber", 0, "Must be 0 when using Orchestrator queues."],
    ])

    ws_assets = wb.create_sheet("Assets")
    write_sheet(ws_assets, HEADER_ASSET, [
        ["CredentialSap", "cred_sap", "SAP", "SAP credential."],
    ])

    ws_env = wb.create_sheet("Environments")
    write_sheet(ws_env, HEADER_STANDARD, [
        ["BaseUrl",     "https://uat.example.com", "API base URL."],
        ["Environment", "UAT",                     "Deployment environment."],
        ["Timeout",     30,                        "HTTP timeout in seconds."],
    ])

    ws_features = wb.create_sheet("Features")
    write_sheet(ws_features, HEADER_STANDARD, [
        ["EnableNotifications", True,  "Send email notifications on completion."],
        ["EnableDryRun",        False, "Skip writes when true."],
        ["MaxParallelJobs",     4,     "Number of parallel job slots."],
    ])

    wb.save(OUTPUT_DIR / "Config_MultiSheet.xlsx")
    print("Created Config_MultiSheet.xlsx")


# ---------------------------------------------------------------------------
# Config_CustomSheets.xlsx — no standard sheet names, tests mapper with any name
# ---------------------------------------------------------------------------
def make_custom_sheets():
    wb = openpyxl.Workbook()

    ws_db = wb.active
    ws_db.title = "Database"
    write_sheet(ws_db, HEADER_STANDARD, [
        ["ConnectionString", "Server=db;Database=app;", "ADO.NET connection string."],
        ["CommandTimeout",   30,                        "Query timeout in seconds."],
        ["MaxPoolSize",      10,                        "Connection pool size."],
    ])

    ws_smtp = wb.create_sheet("Smtp")
    write_sheet(ws_smtp, HEADER_STANDARD, [
        ["Host",        "smtp.example.com", "SMTP server hostname."],
        ["Port",        587,                "SMTP port."],
        ["UseSsl",      True,               "Enable TLS."],
        ["FromAddress", "bot@example.com",  "Sender address."],
    ])

    ws_creds = wb.create_sheet("Credentials")
    write_sheet(ws_creds, HEADER_ASSET, [
        ["SmtpCredential",   "cred_smtp",   "Shared", "SMTP login credential."],
        ["DatabasePassword", "cred_db_pwd", "DB",     "Database user password."],
    ])

    wb.save(OUTPUT_DIR / "Config_CustomSheets.xlsx")
    print("Created Config_CustomSheets.xlsx")


# ---------------------------------------------------------------------------
# Config_BadHeader.xlsx — missing/empty header rows for error handling (#21)
# ---------------------------------------------------------------------------
def make_bad_header():
    wb = openpyxl.Workbook()

    # Sheet 1: completely empty — triggers "missing or empty header row" warning
    ws_empty = wb.active
    ws_empty.title = "EmptySheet"

    # Sheet 2: blank first row — triggers "missing or empty header row" warning
    ws_noheader = wb.create_sheet("NoHeader")
    ws_noheader.append([None, None, None])
    ws_noheader.append(["SomeKey", "SomeValue", "Some description"])

    # Sheet 3: valid — should still generate correctly alongside the bad ones
    ws_settings = wb.create_sheet("Settings")
    write_sheet(ws_settings, HEADER_STANDARD, [
        ["QueueName", "TestQueue", "A valid setting."],
    ])

    wb.save(OUTPUT_DIR / "Config_BadHeader.xlsx")
    print("Created Config_BadHeader.xlsx")


# ---------------------------------------------------------------------------
# Config_Test.xlsx — richest Settings+Constants (all types), empty Assets
# Used as Data/Config.xlsx in the REFramework test fixtures (#32).
# Empty Assets sheet avoids GetRobotAsset calls in CI (no Orchestrator).
# ---------------------------------------------------------------------------
def make_config_test():
    wb = openpyxl.Workbook()

    ws_settings = wb.active
    ws_settings.title = "Settings"
    write_sheet(ws_settings, HEADER_STANDARD, [
        ["FeatureName",       "TypesDemo",                          "string"],
        ["MaxItems",          42,                                   "int"],
        ["Threshold",         3.14,                                 "double"],
        ["IsEnabled",         True,                                 "bool"],
        ["CutoffDate",        datetime.date(2025, 12, 31),          "DateOnly — date only, time is 00:00:00"],
        ["ScheduledAt",       datetime.datetime(2025, 6, 15, 9, 30),"DateTime — has time component"],
        ["DailyRunTime",      datetime.time(8, 0, 0),               "TimeOnly — time only, no date"],
    ])

    ws_constants = wb.create_sheet("Constants")
    write_sheet(ws_constants, HEADER_STANDARD, [
        ["Pi",                3.14159,                              "double — mathematical constant"],
        ["MaxRetryNumber",    0,                                    "int"],
        ["StrictMode",        False,                                "bool"],
        ["ExpiresOn",         datetime.date(2026, 1, 1),            "DateOnly"],
        ["CreatedAt",         datetime.datetime(2024, 3, 1, 12, 0), "DateTime"],
        ["WindowOpen",        datetime.time(9, 0, 0),               "TimeOnly"],
        ["WindowClose",       datetime.time(17, 30, 0),             "TimeOnly"],
    ])

    ws_assets = wb.create_sheet("Assets")
    write_sheet(ws_assets, HEADER_ASSET, [])  # empty — no GetRobotAsset calls in CI

    wb.save(OUTPUT_DIR / "Config_Test.xlsx")
    print("Created Config_Test.xlsx")


# ---------------------------------------------------------------------------
# Config_TypedAssets.xlsx — two asset sheets; one with optional ValueType col
# Sheet 1 "Assets":      5-column format, some rows have ValueType defined
# Sheet 2 "Credentials": 4-column format (no ValueType), all default to object
# Designed for manual testing in UiPath Studio with real Orchestrator assets.
# ---------------------------------------------------------------------------
def make_typed_assets():
    wb = openpyxl.Workbook()

    # Sheet 1: Assets — 5-column format with optional ValueType
    # Folder convention: ConFigTree/Test in Orchestrator
    # Note: credentials are never here — REFramework uses GetRobotCredential separately.
    ws_assets = wb.active
    ws_assets.title = "Assets"
    write_sheet(ws_assets, HEADER_ASSET_TYPED, [
        # Name              Asset name (Orchestrator)        Folder              Description                    ValueType
        ["QueueName",       "cfgtree_queue_name",            "ConFigTree/Test",  "Input queue name.",           "string"],
        ["MaxItemsPerRun",  "cfgtree_max_items_per_run",     "ConFigTree/Test",  "Max items to process.",       "int"],
        ["StrictMode",      "cfgtree_strict_mode",           "ConFigTree/Test",  "Enable strict processing.",   "bool"],
        ["ApiEndpoint",     "cfgtree_api_endpoint",          "ConFigTree/Test",  "REST API base URL.",          "string"],
        ["GenericValue",    "cfgtree_generic_value",         "ConFigTree/Test",  "Untyped — no ValueType.",     None],
    ])

    # Sheet 2: Endpoints — 4-column classic format, no ValueType column (all default to object)
    ws_endpoints = wb.create_sheet("Endpoints")
    write_sheet(ws_endpoints, HEADER_ASSET, [
        # Name              Asset name (Orchestrator)        Folder              Description
        ["BaseUrl",         "cfgtree_base_url",              "ConFigTree/Test",  "REST API base URL."],
        ["OrchestratorFolder", "cfgtree_orch_folder",        "ConFigTree/Test",  "Orchestrator folder path for queue operations."],
    ])

    wb.save(OUTPUT_DIR / "Config_TypedAssets.xlsx")
    print("Created Config_TypedAssets.xlsx")


# ---------------------------------------------------------------------------
# Config_Everything.xlsx — documentation master fixture
# Settings + Constants (REFramework required), 2 extra config sheets,
# 2 asset sheets (one typed, one classic), 2 hidden underscore-prefixed sheets.
# All supported C# types appear in Settings and Constants.
# ---------------------------------------------------------------------------
def make_reference():
    wb = openpyxl.Workbook()

    # --- Hidden sheets first (underscore-prefixed, excluded from code generation) ---

    ws_meta = wb.active
    ws_meta.title = "_Meta"
    write_sheet(ws_meta, ["Key", "Value"], [
        ["Author",        "ConFigTree"],
        ["Version",       "1.0.0"],
        ["Description",   "Documentation master fixture — all types, all sheet kinds."],
        ["GeneratedBy",   "generate_fixtures.py"],
    ])

    # --- Config sheets ---

    ws_settings = wb.create_sheet("Settings")
    write_sheet(ws_settings, HEADER_STANDARD_TYPED, [
        # Name                     Value                                 Description                               DataType
        ["OrchestratorQueueName",  "everything_input_queue",            "string — Orchestrator input queue name.", None],
        ["MaxItemsPerRun",         100,                                  "int — Maximum items to process per run.", None],
        ["Threshold",              3.14,                                 "double — Processing threshold.",          None],
        ["SampleRate",             4,                                    "double forced via DataType (cell is int).","double"],
        ["IsEnabled",              True,                                 "bool — Master feature toggle.",           None],
        ["CutoffDate",             datetime.date(2025, 12, 31),          "DateOnly — Last valid processing date.", None],
        ["ScheduledAt",            datetime.datetime(2025, 6, 15, 9, 30),"DateTime — Next scheduled run.",         None],
        ["DailyRunTime",           datetime.time(8, 0, 0),               "TimeOnly — Daily start time.",           None],
    ])

    ws_constants = wb.create_sheet("Constants")
    write_sheet(ws_constants, HEADER_STANDARD, [
        ["MaxRetryNumber",              0,                                    "int — Must be 0 when using Orchestrator queues."],
        ["MaxConsecutiveSystemExceptions", 3,                                 "int — Stop job after this many consecutive errors."],
        ["Pi",                          3.14159,                              "double — Mathematical constant."],
        ["StrictMode",                  False,                                "bool — Disable for permissive validation."],
        ["ExpiresOn",                   datetime.date(2026, 1, 1),            "DateOnly — Config expiry date."],
        ["CreatedAt",                   datetime.datetime(2024, 3, 1, 12, 0), "DateTime — Config creation timestamp."],
        ["WindowOpen",                  datetime.time(9, 0, 0),               "TimeOnly — Processing window start."],
        ["WindowClose",                 datetime.time(17, 30, 0),             "TimeOnly — Processing window end."],
    ])

    ws_environments = wb.create_sheet("Environments")
    write_sheet(ws_environments, HEADER_STANDARD, [
        ["BaseUrl",        "https://uat.example.com",  "string — REST API base URL."],
        ["Environment",    "UAT",                      "string — Deployment environment name."],
        ["Timeout",        30,                         "int — HTTP request timeout in seconds."],
        ["RetryDelay",     2.5,                        "double — Delay between retries in seconds."],
    ])

    ws_features = wb.create_sheet("Features")
    write_sheet(ws_features, HEADER_STANDARD, [
        ["EnableNotifications",  True,    "bool — Send email on job completion."],
        ["EnableDryRun",         False,   "bool — Skip writes when true."],
        ["MaxParallelJobs",      4,       "int — Parallel job slot limit."],
        ["FeatureLabel",         "beta",  "string — Deployment stage label."],
    ])

    # --- Asset sheets ---

    ws_assets = wb.create_sheet("Assets")
    write_sheet(ws_assets, HEADER_ASSET_TYPED, [
        # Name              Asset name (Orchestrator)      Folder       Description                      ValueType
        ["QueueName",       "cfgtree_queue_name",          "CPMForge",  "Input queue name.",             "string"],
        ["MaxItemsPerRun",  "cfgtree_max_items_per_run",   "CPMForge",  "Maximum items to process.",     "int"],
        ["StrictMode",      "cfgtree_strict_mode",         "CPMForge",  "Strict processing toggle.",     "bool"],
        ["GenericValue",    "cfgtree_generic_value",       "CPMForge",  "Untyped fallback asset.",       None],
    ])

    ws_connections = wb.create_sheet("Connections")
    write_sheet(ws_connections, HEADER_ASSET_TYPED, [
        # Name                  Asset name (Orchestrator)      Folder       Description                       ValueType
        ["ApiEndpoint",         "cfgtree_api_endpoint",        "CPMForge",  "REST API endpoint URL.",          "string"],
        ["BaseUrl",             "cfgtree_base_url",            "CPMForge",  "Service base URL.",               "string"],
        ["OrchestratorFolder",  "cfgtree_orch_folder",         "CPMForge",  "Orchestrator folder path.",       "string"],
    ])

    # --- Hidden sheet: _Notes ---

    ws_notes = wb.create_sheet("_Notes")
    write_sheet(ws_notes, ["Section", "Note"], [
        ["Settings",      "REFramework required — must always be present."],
        ["Constants",     "REFramework required — must always be present."],
        ["Environments",  "Extra config sheet example."],
        ["Features",      "Extra config sheet example."],
        ["Assets",        "5-column typed asset sheet — process runtime values."],
        ["Connections",   "5-column typed asset sheet — endpoint/URL values."],
        ["_Meta",         "Hidden — excluded from code generation (underscore prefix)."],
        ["_Notes",        "Hidden — excluded from code generation (underscore prefix)."],
    ])

    wb.save(OUTPUT_DIR / "Config_Reference.xlsx")
    print("Created Config_Reference.xlsx")


# ---------------------------------------------------------------------------
# Config_ValueTypeOffset.xlsx — regression fixture for #59
# ValueType column is in position 5 (0-indexed), NOT position 4.
# An extra "Tags" column sits at position 4, so row[4] reads "Tags" values.
# With the fix (header.findIndex), types are resolved correctly.
# Without the fix (row[4]), all assets fall back to OrchestratorAsset<object>.
# ---------------------------------------------------------------------------
HEADER_ASSET_OFFSET = ["Name", "Asset", "OrchestratorAssetFolder", "Description", "Tags", "ValueType"]


def make_valuetype_offset():
    wb = openpyxl.Workbook()

    ws_assets = wb.active
    ws_assets.title = "Assets"
    write_sheet(ws_assets, HEADER_ASSET_OFFSET, [
        # Name           Asset name             Folder             Description          Tags        ValueType
        ["QueueName",    "offset_queue",        "ConFigTree/Test", "Queue name.",       "required", "string"],
        ["MaxItems",     "offset_max_items",    "ConFigTree/Test", "Upper bound.",      "required", "int"],
        ["StrictMode",   "offset_strict_mode",  "ConFigTree/Test", "Enable strict.",    "optional", "bool"],
        ["GenericValue", "offset_generic",      "ConFigTree/Test", "Untyped fallback.", "optional", None],
    ])

    wb.save(OUTPUT_DIR / "Config_ValueTypeOffset.xlsx")
    print("Created Config_ValueTypeOffset.xlsx")


# ---------------------------------------------------------------------------
# Config_TargetType.xlsx — regression fixture for #79
# Sheet "SAP" has a _TargetType directive row pointing to an external class.
# Generator should emit ToSapConfig() mapping method; directive row must not
# appear as a property.
# ---------------------------------------------------------------------------
def make_target_type():
    wb = openpyxl.Workbook()

    ws_sap = wb.active
    ws_sap.title = "SAP"
    write_sheet(ws_sap, HEADER_STANDARD, [
        ["Host",   "sap.example.com", "SAP application server hostname."],
        ["Port",   3200,              "SAP system number (integer)."],
        ["UseTLS", True,              "Enable TLS for RFC connection."],
        ["_TargetType", "CPMForge.SAP.SapConfig", None],
    ])

    wb.save(OUTPUT_DIR / "Config_TargetType.xlsx")
    print("Created Config_TargetType.xlsx")


# ---------------------------------------------------------------------------
# Config_CredentialRef.xlsx — fixture for #80
# DataType=credential → string property + two companion getters (Folder, Name)
# ---------------------------------------------------------------------------
def make_credential_ref():
    wb = openpyxl.Workbook()

    ws_sap = wb.active
    ws_sap.title = "SAP"
    write_sheet(ws_sap, HEADER_STANDARD_TYPED, [
        # Name               Value                   Description                                                             DataType
        ["Host",             "sap.example.com",      "SAP application server hostname.",                                     None],
        ["Port",             3200,                    "SAP system number.",                                                   None],
        ["CredentialAsset",  "Default/cred_SAP01",   "Orchestrator credential asset name. When set, Username and Password are ignored.", "credential"],
    ])

    wb.save(OUTPUT_DIR / "Config_CredentialRef.xlsx")
    print("Created Config_CredentialRef.xlsx")


# ---------------------------------------------------------------------------
# Config_AssetRef.xlsx — fixture for #81
# DataType=asset → identical companion getters as DataType=credential (#80)
# ---------------------------------------------------------------------------
def make_asset_ref():
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "Queue"
    write_sheet(ws, HEADER_STANDARD_TYPED, [
        # Name           Value                    Description                                   DataType
        ["MaxItems",     50,                      "Maximum items per run.",                     None],
        ["QueueAsset",   "Default/cfgtree_queue", "Orchestrator asset holding the queue name.", "asset"],
    ])

    wb.save(OUTPUT_DIR / "Config_AssetRef.xlsx")
    print("Created Config_AssetRef.xlsx")


# ---------------------------------------------------------------------------
# Config_DataType.xlsx — regression fixture for #85
# Verifies DataType column override logic:
#   - int cell forced to double via DataType="double"
#   - empty DataType → inferred type (int)
#   - unrecognised DataType → inferred fallback (string)
#   - credential and asset sentinels in the same sheet
# ---------------------------------------------------------------------------
def make_datatype():
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "Overrides"
    write_sheet(ws, HEADER_STANDARD_TYPED, [
        # Name            Value                    Description                              DataType
        ["SampleRate",    4,                       "int cell forced to double.",            "double"],
        ["MaxRetries",    3,                       "int inferred (empty DataType).",        None],
        ["Label",         "beta",                  "string inferred (bogus DataType).",     "bogus"],
        ["CredentialSap", "Default/cred_sap_dt",  "Credential ref sentinel.",              "credential"],
        ["QueueRef",      "Default/cfgtree_dt",   "Asset ref sentinel.",                   "asset"],
    ])

    wb.save(OUTPUT_DIR / "Config_DataType.xlsx")
    print("Created Config_DataType.xlsx")


# ---------------------------------------------------------------------------
# Config_Combined.xlsx — regression fixture for #86
# _TargetType + DataType=credential + DataType=asset in the same sheet.
# Golden confirms: companion getters emitted, ToSapConfig() maps base
# string props only (not companion getters).
# ---------------------------------------------------------------------------
def make_combined():
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "SAP"
    write_sheet(ws, HEADER_STANDARD_TYPED, [
        # Name               Value                           Description                                                              DataType
        ["Host",             "sap.example.com",             "SAP application server hostname.",                                      None],
        ["Port",             3200,                          "SAP system number.",                                                    None],
        ["CredentialAsset",  "Default/cred_SAP01",          "Orchestrator credential. When set, Username and Password are ignored.", "credential"],
        ["QueueAsset",       "Default/cfgtree_queue",       "Orchestrator asset holding the queue name.",                           "asset"],
        ["_TargetType",      "CPMForge.SAP.SapConfig", None,                                                               None],
    ])

    wb.save(OUTPUT_DIR / "Config_Combined.xlsx")
    print("Created Config_Combined.xlsx")


# ---------------------------------------------------------------------------
# Config_AssetRobustness.xlsx — regression fixture for #60
#
# Demonstrates that the runtime xlsx column structure for asset sheets is
# irrelevant: Load() always skips asset sheets (they are populated via
# GetRobotAsset in the generated XAML snippet, never from DataTable).
#
# Generation-time xlsx: TypedAssets sheet has 5 columns (ValueType present).
# Runtime xlsx may have 0..N columns — it does not matter.
# The golden proves this by showing TypedAssets absent from Load().
#
# Cases covered:
#   1. Asset sheet with no ValueType column (4 cols) → object? properties
#   2. Asset sheet with ValueType column but all cells empty → object? properties
#   3. Extra column beyond ValueType → no crash (header.findIndex handles it)
# ---------------------------------------------------------------------------
HEADER_ASSET_EXTRA = ["Name", "Asset", "OrchestratorAssetFolder", "Description", "ValueType", "Tags"]


def make_asset_robustness():
    wb = openpyxl.Workbook()

    # Standard config sheet — proves Load() is generated for non-asset sheets
    ws_settings = wb.active
    ws_settings.title = "Settings"
    write_sheet(ws_settings, HEADER_STANDARD, [
        ["Environment", "UAT", "Deployment environment."],
    ])

    # Case 1: classic 4-column asset sheet — no ValueType column at all
    ws_classic = wb.create_sheet("ClassicAssets")
    write_sheet(ws_classic, HEADER_ASSET, [
        ["QueueName",   "cfgtree_queue_name",    "ConFigTree/Test", "Input queue name."],
        ["ApiEndpoint", "cfgtree_api_endpoint",  "ConFigTree/Test", "REST API base URL."],
    ])

    # Case 2: 5-column sheet with ValueType column present but all cells empty
    ws_empty_vt = wb.create_sheet("EmptyValueType")
    write_sheet(ws_empty_vt, HEADER_ASSET_TYPED, [
        # Name              Asset                      Folder             Description            ValueType
        ["CredentialSap",   "cred_sap_robust",         "SAP",             "SAP credential.",     None],
        ["CredentialM365",  "cred_m365_robust",        "Shared",          "M365 credential.",    None],
    ])

    # Case 3: 6-column sheet with an extra column after ValueType
    ws_extra = wb.create_sheet("ExtraColumn")
    write_sheet(ws_extra, HEADER_ASSET_EXTRA, [
        # Name          Asset                    Folder             Description          ValueType   Tags
        ["ApiKey",      "cfgtree_api_key",       "ConFigTree/Test", "API key asset.",    "string",   "required"],
        ["MaxItems",    "cfgtree_max_items",     "ConFigTree/Test", "Max items asset.",  "int",      "optional"],
    ])

    wb.save(OUTPUT_DIR / "Config_AssetRobustness.xlsx")
    print("Created Config_AssetRobustness.xlsx")


# ---------------------------------------------------------------------------
# Config_InvalidNames.xlsx — exercises the C# identifier sanitizer:
#   - Sheet whose PascalCased name starts with a digit ("2024-Q1")
#   - Sheet with a character that survives PascalCase ("Bad Sheet!")
#   - Sheet whose PascalCased name is empty ("!!!")   → "IdentifierConfig"
#   - Two sheets that PascalCase to the same identifier collide
#     ("my-section" and "My Section" both → "MySectionConfig")
#   - Properties with special chars and a within-sheet name collision
#
# NB: Excel prevents sheet names that differ only by case (it would rename the
#     second one), so collision-by-casing can only be tested at the property
#     level or via two sheet names that PascalCase to the same identifier.
# ---------------------------------------------------------------------------
def make_invalid_names():
    wb = openpyxl.Workbook()

    ws_digit = wb.active
    ws_digit.title = "2024-Q1"
    write_sheet(ws_digit, HEADER_STANDARD, [
        ["MaxItems",  10,    "Max items per quarter."],
        ["LogPrefix", "Q1",  "Prefix for log messages."],
    ])

    ws_bad = wb.create_sheet("Bad Sheet!")
    write_sheet(ws_bad, HEADER_STANDARD, [
        ["Feature#1",   "on",  "First feature toggle."],
        ["user.email",  "",    "Contact email."],
        ["user-email",  "",    "Collides with user.email after PascalCase."],
    ])

    ws_empty = wb.create_sheet("!!!")
    write_sheet(ws_empty, HEADER_STANDARD, [
        ["Note", "placeholder", "Survives despite degenerate sheet name."],
    ])

    ws_a = wb.create_sheet("my-section")
    write_sheet(ws_a, HEADER_STANDARD, [
        ["A", "1", ""],
    ])

    ws_b = wb.create_sheet("My Section")
    write_sheet(ws_b, HEADER_STANDARD, [
        ["B", "2", ""],
    ])

    wb.save(OUTPUT_DIR / "Config_InvalidNames.xlsx")
    print("Created Config_InvalidNames.xlsx")


if __name__ == "__main__":
    make_basic()
    make_types()
    make_types_toml()
    make_assets()
    make_multi_sheet()
    make_custom_sheets()
    make_bad_header()
    make_config_test()
    make_typed_assets()
    make_reference()
    make_valuetype_offset()
    make_target_type()
    make_credential_ref()
    make_asset_ref()
    make_datatype()
    make_combined()
    make_asset_robustness()
    make_invalid_names()
    print("Done.")
